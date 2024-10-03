import plotly.graph_objects as go
import plotly.io as pio
def plot_df(df, *columns, fig=None, show=False):
    colors = ['red', 'blue', 'green', 'orange', 'purple', 'cyan', 'magenta']

    if fig is None:
        fig = go.Figure()

    if 'ix' not in df.columns:
        df['ix'] = range(1, len(df) + 1)

    for i, column in enumerate(columns):
        # Create hovertext that shows both the label and the value
        hovertext = [f"{column}: {val}<br>{index.strftime('%Y-%m-%d %H:%M:%S')}"
                     for val, index in zip(df[column], df.index)]

        fig.add_trace(go.Scatter(
            x=df['ix'],
            y=df[column],
            mode='lines',
            name=f'{column}',
            line=dict(color=colors[i % len(colors)]),
            hovertext=hovertext,  # Use the combined hovertext with labels and values
            hoverinfo='text',  # 'text' will show custom hovertext
        ))

    fig.update_layout(
        title="",
        xaxis_title="Index",
        yaxis_title="Values",
        plot_bgcolor='rgba(30, 30, 30, 1)',  # Dark gray background
        paper_bgcolor='rgba(30, 30, 30, 1)',  # Same for outer background
        font=dict(color='white'),  # White font for better contrast
        margin=dict(l=20, r=20, t=20, b=20),  # Keep some margin for better framing
        xaxis=dict(
            showgrid=False,
            zeroline=False,  # Removing the centerline to clean up
            color='white',  # White labels for dark background
        ),
        yaxis=dict(
            showgrid=False,
            zeroline=False,
            color='white',
        ),
        hovermode='x',  # Enable hover on x-axis for better usability
    )

    if show:
        fig.show()

    filename_columns = "_".join(columns)
    filename = f"{filename_columns}.html"
    # pio.write_html(fig, file=filename, auto_open=False)

    return fig


def save_plot(fig, filename):
    pio.write_html(fig, file=filename, auto_open=False)

def draw_horizontal_line(fig, y, x0, x1, color='Red'):
    fig.add_shape(
        type="line",
        x0=x0, x1=x1,
        y0=y, y1=y,
        line=dict(color=color, width=2, dash="dash"),
        name=f"{y} Line"
    )
    return fig


import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def save_excel_trades(file_path):
    df = pd.read_excel(file_path)

    # Sort by timestamp first, then by symbol
    df.sort_values(by=['timestamp', 'symbol'], inplace=True)

    # Reorder the columns (timestamp first, then symbol)
    columns = ['timestamp', 'symbol'] + [col for col in df.columns if col not in ['timestamp', 'symbol']]
    df = df[columns]


    # Create a new Workbook object
    wb = Workbook()
    ws = wb.active

    # Write the headers with bold formatting and padding
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = alignment_center
        ws.row_dimensions[1].height = 20  # Increase header row height for padding

    # Write the sorted data into the worksheet row by row, skipping the index
    for row in df.itertuples(index=False, name=None):
        ws.append(row)

    # Freeze the top row (first row)
    ws.freeze_panes = "A2"

    # Define fills for gray, white, and BANKNIFTY highlight
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    banknifty_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight

    # Initialize variables for row coloring
    previous_timestamp = None
    current_fill = white_fill

    # Loop through each row, starting from the second row (because the first row is headers)
    for row in range(2, ws.max_row + 1):
        # Get the timestamp and symbol for the current row
        current_timestamp = ws[f'A{row}'].value
        current_symbol = ws[f'B{row}'].value

        # If the timestamp changes, alternate the fill color
        if current_timestamp != previous_timestamp:
            current_fill = gray_fill if current_fill == white_fill else white_fill

        # Apply the current fill color to the entire row
        for col in range(1, ws.max_column + 1):
            if current_symbol == "BANKNIFTY":
                ws.cell(row=row, column=col).fill = banknifty_fill  # Highlight BANKNIFTY
            else:
                ws.cell(row=row, column=col).fill = current_fill  # Apply alternating gray/white blocks for timestamps

        # Update the previous timestamp to the current one
        previous_timestamp = current_timestamp

    # Auto-size all columns to fit the content
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2  # Add a little extra space

    # Save the formatted Excel file
    wb.save(file_path)

# Function to check if only Entry or Exit is "Yes" and others are "No"
def is_special_row(row):
    entry_exit_yes = (row['ENTRY'] == 'Yes') or (row['EXIT'] == 'Yes')
    others_no = all(val == 'No' for key, val in row.items() if key not in ['ENTRY', 'EXIT'])
    return entry_exit_yes and others_no

def save_excel_trade_execution(input_file):
    df = pd.read_excel(input_file)

    # Define pleasant color fills for "Yes", "No", and special row conditions
    green_fill = PatternFill(start_color='B7E1CD', end_color='B7E1CD', fill_type='solid')  # Pleasant Green for Yes
    red_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')    # Soft Red for No
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Pastel Yellow for special rows

    # Load the workbook and select the active sheet
    workbook = load_workbook(input_file)
    sheet = workbook.active

    # Freeze the header row
    sheet.freeze_panes = "B2"  # This freezes the header (first row and index column)

    # Set column width for better readability, including the index column
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2  # Add extra padding to the width
        sheet.column_dimensions[column].width = adjusted_width

    # Ensure the index column (typically column A) is padded
    sheet.column_dimensions['A'].width = 15  # Set a wider width for the index column

    # Iterate over each row and apply the appropriate colors
    for idx, row in df.iterrows():
        excel_row = idx + 2  # Adjust for Excel's 1-based indexing and header row

        # Check if this row meets the special condition
        if is_special_row(row):
            for cell in sheet[excel_row]:
                cell.fill = yellow_fill
        else:
            for col_idx, value in enumerate(row):
                cell = sheet.cell(row=excel_row, column=col_idx + 1)
                if value == 'Yes':
                    cell.fill = green_fill
                elif value == 'No':
                    cell.fill = red_fill

    # Save the updated Excel file
    workbook.save(input_file)