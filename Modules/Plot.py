import plotly.graph_objects as go


def plot_df(df, *columns, **kwargs):
    colors = [
        '#4C72B0',  # Soft blue
        '#55A868',  # Soft green
        '#C44E52',  # Soft red
        '#8172B3',  # Soft purple
        '#CCB974',  # Soft yellow
        '#64B5CD',  # Soft cyan
        '#D3A965',  # Warm tan
        '#8C8C8C',  # Soft gray
        '#E17C05',  # Muted orange
        '#76B7B2',  # Muted teal
        '#A8786E',  # Muted brown
        '#F28E2B',  # Soft peach
        '#B07AA1',  # Soft mauve
        '#59A14F',  # Soft lime green
        '#EDC948',  # Soft mustard
    ]
    import pandas as pd
    df = df.where(pd.notna(df), None)
    # Ensure 'ix' column exists for indexing if necessary
    if 'ix' not in df.columns:
        df['ix'] = range(1, len(df) + 1)

    fig = go.Figure()

    if isinstance(df.index, pd.DatetimeIndex):
        label_x = "Timestamp"
    else:
        label_x = "Underlying Price"
    
    label_y = "Values"
    
    if 'x' in kwargs.keys():
        label_x = kwargs['x']

    if 'y' in kwargs.keys():
        label_y = kwargs['y']
    
    hovertext = [
        f"{label_x}: {index}" for index in df.index
    ]
    
    fig.add_trace(go.Scatter(
        x=df['ix'],
        y=[0]*len(df),
        mode='text',
        name=label_x,
        line=dict(color=colors[0 % len(colors)]),
        hoverinfo='text',  # Display custom hovertext
        hovertext=hovertext  # Custom hover text with timestamp and value
    ))
    for i, column in enumerate(columns):
        # Create custom hovertext to show timestamp and column value
        hovertext = [
            f"{column}: {value:.2f}" if pd.notna(value) else ""
            for index, value in zip(df.index, df[column])
        ]
        
        fig.add_trace(go.Scatter(
            x=df['ix'],
            y=df[column],
            mode='lines',
            name=column,
            line=dict(color=colors[i % len(colors)]),
            hoverinfo='text',  # Display custom hovertext
            hovertext=hovertext  # Custom hover text with timestamp and value
        ))

    fig.update_layout(
        title="Plot: " + ", ".join(df.columns),
        xaxis_title=f"{label_x} (x)",
        yaxis_title=f"{label_y} (y)",
        plot_bgcolor='rgba(30, 30, 30, 1)',  # Dark gray background
        paper_bgcolor='rgba(30, 30, 30, 1)',  # Outer background
        font=dict(color='white'),  # White font for better contrast
        margin=dict(l=20, r=20, t=20, b=20),  # Keep some margin for better framing
        xaxis=dict(
            showgrid=False,
            zeroline=False,
            color='white'
        ),
        yaxis=dict(
            showgrid=False,
            zeroline=False,
            color='white'
        ),
        hovermode='x unified'  # Show a single hover box on the x-axis with all trace info
    )

    return fig


def draw_horizontal_line(fig, y, x0, x1, color='Red', width=0.8):
    import plotly.graph_objects as go
    fig.add_shape(
        type="line",
        x0=x0, x1=x1,
        y0=y, y1=y,
        line=dict(color=color, width=width, dash="dash"),
        name=f"{y} Line"
    )
    return fig


def save_plot(fig, filename):
    import plotly.io as pio
    pio.write_html(fig, file=filename, auto_open=False)


def save_df_to_excel(data_dict, file_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import pandas as pd
    from enum import Enum

    # Define styles and formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")  # Muted blue for headers
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Define color palette in valid aRGB format
    colors = [
    'FFB0C4DE',  # Light Steel Blue
    'FF98FB98',  # Pale Green
    'FFFFB6C1',  # Light Pink
    'FFB19CD9',  # Soft Lavender
    'FFF5DEB3',  # Wheat
    'FFADD8E6',  # Light Blue
    'FFFFE4B5',  # Moccasin
    'FFD3D3D3',  # Light Gray
    'FFFFEFD5',  # Papaya Whip
    'FFE0FFFF',  # Light Cyan
    'FFF5F5DC',  # Beige
    'FFFFFACD',  # Lemon Chiffon
    'FFE6E6FA',  # Lavender
    'FFF0FFF0',  # Honeydew
    'FFFDF5E6',  # Old Lace
    ]

    light_colors = [
        'FFE6EEF3',  # Lighter Steel Blue
        'FFE8FBE8',  # Lighter Pale Green
        'FFFFE4E9',  # Lighter Pink
        'FFEDE8F5',  # Lighter Lavender
        'FFFFF5E0',  # Lighter Wheat
        'FFEEF6FA',  # Lighter Blue
        'FFFFF7E4',  # Lighter Moccasin
        'FFF0F0F0',  # Lighter Gray
        'FFFFFBEE',  # Lighter Papaya
        'FFEFFDFD',  # Lighter Cyan
        'FFFFFDF6',  # Lighter Beige
        'FFFFFFF2',  # Lighter Lemon
        'FFF9F9FC',  # Lighter Lavender
        'FFF9FFF9',  # Lighter Honeydew
        'FFFDFBF5',  # Lighter Old Lace
    ]

    # Create a new Workbook
    wb = Workbook()

    for sheet_name, df in data_dict.items():
        # Replace NaN with "N/A" and enums with their string values
        df = df.map(lambda x: "N/A" if (pd.isna(x) or x == None) else str(x) if isinstance(x, Enum) else x)

        # Add a new sheet for each DataFrame
        ws = wb.create_sheet(title=sheet_name)

        # Determine the columns associated with each leg
        leg_columns = {}
        for i in range(1, 15):  # Assuming up to 15 legs for simplicity
            leg_columns[f"Leg {i}"] = [
                col for col in df.columns if (f"Leg({i})" in col or f"(Leg {i})" in col)
            ]

        # Write headers with formatting
        ws.cell(row=1, column=1, value="DateTimestamp/ Index")  # Add index header
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).alignment = alignment_center
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).border = thin_border

        for col_num, header in enumerate(df.columns, 2):  # Start from column 2
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = alignment_center
            cell.fill = header_fill
            cell.border = thin_border

        # Write data row-by-row including the index
        for row_idx, (index_value, row) in enumerate(df.iterrows(), start=2):
            # Write the index value in the first column
            index_cell = ws.cell(row=row_idx, column=1, value=index_value)
            index_cell.border = thin_border
            index_cell.alignment = alignment_center

            # Write the rest of the row values
            for col_idx, value in enumerate(row, start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value="N/A" if pd.isna(value) else value)
                cell.border = thin_border
                cell.alignment = alignment_center

                # Determine the column's leg and assign colors
                column_leg = None
                for leg, cols in leg_columns.items():
                    if df.columns[col_idx - 2] in cols:
                        column_leg = leg
                        break

                if column_leg:
                    leg_index = int(column_leg.split()[-1]) - 1  # Get leg index (0-based)
                    base_color = colors[leg_index % len(colors)]
                    light_color = light_colors[leg_index % len(light_colors)]
                    cell.fill = PatternFill(
                        start_color=light_color if row_idx % 2 == 0 else base_color,
                        end_color=light_color if row_idx % 2 == 0 else base_color,
                        fill_type="solid",
                    )
                else:
                    # Default alternating row colors for non-leg columns
                    cell.fill = PatternFill(
                        start_color="FFF2F2F2" if row_idx % 2 == 0 else "FFFFFFFF",
                        end_color="FFF2F2F2" if row_idx % 2 == 0 else "FFFFFFFF",
                        fill_type="solid",
                    )
                    index_cell.fill = PatternFill(
                        start_color="FFF2F2F2" if row_idx % 2 == 0 else "FFFFFFFF",
                        end_color="FFF2F2F2" if row_idx % 2 == 0 else "FFFFFFFF",
                        fill_type="solid",
                    )

        # Freeze the header row and index column
        ws.freeze_panes = "B2"

        # Adjust column widths
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 10)

    # Remove the default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Save the workbook
    wb.save(file_path)





# def plot_df(df, *columns, **kwargs):
#     colors = [
#         '#4C72B0',  # Soft blue
#         '#55A868',  # Soft green
#         '#C44E52',  # Soft red
#         '#8172B3',  # Soft purple
#         '#CCB974',  # Soft yellow
#         '#64B5CD',  # Soft cyan
#         '#D3A965',  # Warm tan
#         '#8C8C8C',  # Soft gray
#         '#E17C05',  # Muted orange
#         '#76B7B2',  # Muted teal
#         '#A8786E',  # Muted brown
#         '#F28E2B',  # Soft peach
#         '#B07AA1',  # Soft mauve
#         '#59A14F',  # Soft lime green
#         '#EDC948',  # Soft mustard
#     ]
#     import pandas as pd
#     df = df.where(pd.notna(df), None)
#     # Ensure 'ix' column exists for indexing if necessary
#     if 'ix' not in df.columns:
#         df['ix'] = range(1, len(df) + 1)

#     fig = go.Figure()

#     if isinstance(df.index, pd.DatetimeIndex):
#         label_x = "Timestamp"
#     else:
#         label_x = "Underlying Price"
    
#     label_y = "Values"
    
#     if 'x' in kwargs.keys():
#         label_x = kwargs['x']

#     if 'y' in kwargs.keys():
#         label_y = kwargs['y']
    
#     # hovertext = [
#     #     f"{label_x}: {index}" for index in df.index
#     # ]
    
#     # fig.add_trace(go.Scatter(
#     #     x=df['ix'],
#     #     y=[0]*len(df),
#     #     mode='text',
#     #     name=label_x,
#     #     line=dict(color=colors[0 % len(colors)]),
#     #     hoverinfo='text',  # Display custom hovertext
#     #     hovertext=hovertext  # Custom hover text with timestamp and value
#     # ))
#     for i, column in enumerate(columns):
#         # Create custom hovertext to show timestamp and column value
#         hovertext = [
#             f"{column}: {value:.2f}" if pd.notna(value) else ""
#             for index, value in zip(df.index, df[column])
#         ]
        
#         fig.add_trace(go.Scatter(
#             x=df.index,
#             y=df[column],
#             mode='lines',
#             name=column,
#             line=dict(color=colors[i % len(colors)]),
#             hoverinfo='text',  # Display custom hovertext
#             customdata=df.index,
#             hovertemplate="%{y:.2f}"  # Dynamically show the column name
#             # hovertext=hovertext,
#         ))

#     # Add a fixed annotation for timestamp
#     fig.update_layout(
#         annotations=[
#             dict(
#                 x=0.5,
#                 y=1.1,
#                 xref="paper",
#                 yref="paper",
#                 text="Hover over the chart to see timestamps",
#                 showarrow=False,
#                 font=dict(size=12, color="white"),
#                 align="center",
#                 bgcolor="rgba(50, 50, 50, 0.8)",
#                 bordercolor="white",
#                 borderwidth=1,
#             )
#         ],
#         title="Plot: " + ", ".join(columns),
#         xaxis_title=label_x,
#         yaxis_title=label_y,
#         plot_bgcolor='rgba(30, 30, 30, 1)',
#         paper_bgcolor='rgba(30, 30, 30, 1)',
#         font=dict(color='white'),
#         margin=dict(l=20, r=20, t=20, b=20),
#         xaxis=dict(
#             showgrid=False,
#             zeroline=False,
#             color='white'
#         ),
#         yaxis=dict(
#             showgrid=False,
#             zeroline=False,
#             color='white'
#         ),
#         hovermode='x unified'
#     )

#     return fig



# def save_df_to_excel(data_dict, file_path):
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
#     from openpyxl.utils import get_column_letter
#     import pandas as pd
#     from enum import Enum

#     # Create a new Workbook
#     wb = Workbook()

#     # Define styles and formatting
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Muted blue for headers
#     row_fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray for even rows
#     row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White for odd rows
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
#     alignment_center = Alignment(horizontal="center", vertical="center")

#     for sheet_name, df in data_dict.items():
#         # Replace NaN with "N/A" and enums with their string values
#         df = df.map(lambda x: "N/A" if (pd.isna(x) or x == None) else str(x) if isinstance(x, Enum) else x)

#         # Add a new sheet for each DataFrame
#         ws = wb.create_sheet(title=sheet_name)
        
#         # Write headers with bold font, center alignment, and extra padding in the header row
#         ws.column_dimensions["A"].width = 15  # Extra width for the index column
#         ws.cell(row=1, column=1, value="DateTimestamp/ Index")  # Add "Index" as the header for the index column
#         ws.cell(row=1, column=1).font = header_font
#         ws.cell(row=1, column=1).alignment = alignment_center
#         ws.cell(row=1, column=1).fill = header_fill
#         ws.cell(row=1, column=1).border = thin_border

#         for col_num, header in enumerate(df.columns, 2):  # Start from column 2 to leave space for the index
#             cell = ws.cell(row=1, column=col_num, value=header)
#             cell.font = header_font
#             cell.alignment = alignment_center
#             cell.fill = header_fill
#             cell.border = thin_border
#             ws.row_dimensions[1].height = 25  # Extra height for header row

#         # Write data row-by-row including the index
#         for row_idx, (index_value, row) in enumerate(df.iterrows(), start=2):
#             # Write the index value in the first column
#             index_cell = ws.cell(row=row_idx, column=1, value=index_value)
#             index_cell.border = thin_border
#             index_cell.alignment = alignment_center
#             index_cell.fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd

#             # Write the rest of the row values
#             for col_idx, value in enumerate(row, start=2):
#                 cell = ws.cell(row=row_idx, column=col_idx, value="N/A" if pd.isna(value) else value)
#                 cell.border = thin_border
#                 cell.alignment = alignment_center
#                 cell.fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd

#         # Freeze the header row and index column
#         ws.freeze_panes = "B2"

#         # Adjust column widths based on maximum cell lengths in each column
#         for col_idx, column_cells in enumerate(ws.columns, 1):
#             max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
#             ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 10)

#     # Remove the default sheet if it exists
#     if 'Sheet' in wb.sheetnames:
#         wb.remove(wb['Sheet'])

#     # Save the workbook
#     wb.save(file_path)
